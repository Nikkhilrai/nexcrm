import re

from django.db.models import Prefetch, ProtectedError
from django.http import HttpResponse
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import CursorPagination
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.permissions import IsAdmin, IsAdminOrReadUpdate

from .filters import LeadFilter
from .models import Contact, Interaction, Lead, PackageTier, StatusHistory, SubPipeline
from .serializers import (
    ContactSerializer,
    InteractionSerializer,
    LeadDetailSerializer,
    LeadListSerializer,
    PackageTierSerializer,
    StatusHistorySerializer,
    SubPipelineSerializer,
)


class LeadCursorPagination(CursorPagination):
    """Cursor pagination on -created_at for the infinite-scroll table.

    Page size is fixed at 25 — the home table loads 25 rows per scroll.
    """

    page_size = 25
    ordering = "-created_at"


class LeadViewSet(viewsets.ModelViewSet):
    """Full CRUD on leads.

    - GET / POST / PATCH / PUT — any authenticated user
    - DELETE — admin only (enforced by IsAdminOrReadUpdate)
    - List uses LeadListSerializer (compact); detail uses LeadDetailSerializer (full)
    """

    permission_classes = (IsAdminOrReadUpdate,)
    pagination_class = LeadCursorPagination
    filterset_class = LeadFilter
    search_fields = ("full_name", "phone", "email", "company")
    ordering_fields = ("created_at", "next_followup_at", "updated_at")

    def get_queryset(self):
        # Prefetch status_history descending by changed_at so the list
        # serializer's `latest_status_note` field can pull the newest row
        # via `obj.status_history.all()[0]` without an extra query per lead.
        # The model already orders by -changed_at; we restate it here so the
        # intent is local to the queryset.
        history_qs = StatusHistory.objects.select_related("changed_by").order_by(
            "-changed_at"
        )
        return Lead.objects.select_related(
            "event_interest",
            "assigned_to",
            "created_by",
            "package_tier",
            "sub_pipeline",
        ).prefetch_related(Prefetch("status_history", queryset=history_qs))

    def get_serializer_class(self):
        if self.action == "list":
            return LeadListSerializer
        return LeadDetailSerializer

    @action(detail=False, methods=["get"], url_path="by-phone")
    def by_phone(self, request):
        """GET /api/leads/by-phone/?phone=+91...&exclude=<uuid>

        Powers the edit-page sidebar — "past leads from same phone number".
        Returns ALL leads matching the phone (no pagination — there's
        rarely more than a handful per number).
        """
        phone = request.query_params.get("phone", "").strip()
        if not phone:
            raise ValidationError({"phone": "phone query param is required."})

        qs = self.get_queryset().filter(phone=phone)
        exclude = request.query_params.get("exclude", "").strip()
        if exclude:
            qs = qs.exclude(pk=exclude)
        # Most recent first so the sidebar reads top-down.
        qs = qs.order_by("-created_at")
        return Response(LeadListSerializer(qs, many=True).data)

    @action(
        detail=True,
        methods=["get", "post"],
        url_path="interactions",
    )
    def interactions(self, request, pk=None):
        """GET → list interactions for this lead.
        POST → log a new interaction (lead + user are filled server-side).
        """
        lead = self.get_object()

        if request.method == "GET":
            qs = lead.interactions.select_related("user").all()
            return Response(InteractionSerializer(qs, many=True).data)

        # POST
        serializer = InteractionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        interaction = Interaction.objects.create(
            lead=lead,
            user=request.user,
            **serializer.validated_data,
        )
        return Response(
            InteractionSerializer(interaction).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["get"], url_path="status-history")
    def status_history(self, request, pk=None):
        """GET → status timeline for this lead. Read-only; rows are
        auto-created by the post_save signal on Lead.
        """
        lead = self.get_object()
        qs = lead.status_history.select_related("changed_by").all()
        return Response(StatusHistorySerializer(qs, many=True).data)


class ContactCursorPagination(CursorPagination):
    """Cursor pagination for /api/contacts/ list — by name for predictable
    typeahead ordering. Page size matches the picker's max display.
    """

    page_size = 25
    ordering = "full_name"


class ContactViewSet(viewsets.ModelViewSet):
    """CRUD on the contacts database.

    - GET (list/retrieve) — any authenticated user (LeadForm picker uses search).
    - POST / PATCH / PUT / DELETE — admin only.
    - `?search=<q>` matches against full_name, phone, email, company.
    """

    serializer_class = ContactSerializer
    pagination_class = ContactCursorPagination
    search_fields = ("full_name", "phone", "email", "company")

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsAuthenticated()]
        return [IsAdmin()]

    def get_queryset(self):
        return Contact.objects.select_related("created_by").all()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-upload",
        permission_classes=[IsAdmin],
        parser_classes=[MultiPartParser, FormParser],
    )
    def bulk_upload(self, request):
        """POST /api/contacts/bulk-upload/[?dry_run=1] — multipart `.xlsx` upload.

        Expected columns (case-insensitive, in any order):
        Phone (required), Name, Email, Company, Designation, LinkedIn URL, Source.

        Returns {imported: N, skipped: [{row, reason, phone?}], preview: [...]}.
        `preview` echoes the rows that were (or would be, in dry-run) imported
        so the admin can sanity-check before committing.

        When `?dry_run=1` is passed, no DB writes happen — the response shows
        what WOULD be imported. Frontend uses this for the confirm-before-import
        UX.
        """
        upload = request.FILES.get("file")
        if upload is None:
            raise ValidationError({"file": "No file uploaded. Send as multipart form-data with field name 'file'."})
        name = (upload.name or "").lower()
        if not name.endswith(".xlsx"):
            raise ValidationError({"file": "Only .xlsx files are supported."})

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValidationError({"file": "Server is missing the openpyxl dependency."})

        try:
            wb = load_workbook(upload, read_only=True, data_only=True)
        except Exception as e:
            raise ValidationError({"file": f"Couldn't open the workbook: {e}"})

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValidationError({"file": "Workbook is empty."})

        col_map = _resolve_header_columns(header_row)
        if "phone" not in col_map:
            raise ValidationError({"file": "Missing required 'Phone' column in the header row."})

        source_lookup = _build_source_lookup()
        existing_phones = set(Contact.objects.values_list("phone", flat=True))
        seen_in_file: set[str] = set()

        imported_payload = []
        skipped: list[dict] = []
        # Header is row 1; first data row is sheet row 2.
        for offset, row in enumerate(rows_iter, start=2):
            if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue  # blank line — skip silently

            raw_phone = _cell(row, col_map.get("phone"))
            phone = _normalize_phone(raw_phone)
            if not phone:
                skipped.append({"row": offset, "reason": "Missing phone."})
                continue
            if not _looks_like_phone(phone):
                skipped.append({"row": offset, "phone": phone, "reason": "Phone is corrupted (not a real number)."})
                continue
            if phone in seen_in_file:
                skipped.append({"row": offset, "phone": phone, "reason": "Duplicate phone within this file."})
                continue
            if phone in existing_phones:
                skipped.append({"row": offset, "phone": phone, "reason": "Phone already exists in contacts."})
                continue

            # Email and LinkedIn are imported as-is — even if they look junk.
            # Admin can clean them up later from the contact's detail view.
            email = (_cell(row, col_map.get("email")) or "").strip()
            linkedin = (_cell(row, col_map.get("linkedin_url")) or "").strip()

            source_raw = (_cell(row, col_map.get("source")) or "").strip()
            source = source_lookup.get(source_raw.lower(), Contact._meta.get_field("source").default)

            imported_payload.append(
                Contact(
                    phone=phone,
                    full_name=(_cell(row, col_map.get("full_name")) or "").strip()[:200],
                    email=email,
                    company=(_cell(row, col_map.get("company")) or "").strip()[:200],
                    designation=(_cell(row, col_map.get("designation")) or "").strip()[:200],
                    linkedin_url=linkedin,
                    source=source,
                    created_by=request.user,
                )
            )
            seen_in_file.add(phone)

        dry_run = str(request.query_params.get("dry_run", "")).lower() in {"1", "true", "yes"}
        if imported_payload and not dry_run:
            Contact.objects.bulk_create(imported_payload)

        preview = [
            {
                "phone": c.phone,
                "full_name": c.full_name,
                "email": c.email,
                "company": c.company,
                "designation": c.designation,
                "linkedin_url": c.linkedin_url,
                "source": c.source,
            }
            for c in imported_payload
        ]
        return Response(
            {
                "imported": 0 if dry_run else len(imported_payload),
                "would_import": len(imported_payload),
                "dry_run": dry_run,
                "skipped": skipped,
                "preview": preview,
            },
            status=status.HTTP_200_OK,
        )

    @action(
        detail=False,
        methods=["get"],
        url_path="template",
        permission_classes=[IsAdmin],
    )
    def template(self, request):
        """GET /api/contacts/template/ — downloadable .xlsx with the canonical
        7 headers plus 2 sample rows so the admin sees the expected format.
        """
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(["Phone", "Name", "Email", "Company", "Designation", "LinkedIn URL", "Source"])
        ws.append([
            "+919876543210",
            "Aanya Sharma",
            "aanya@example.com",
            "Sharma & Co.",
            "Managing Partner",
            "https://www.linkedin.com/in/aanya-sharma/",
            "LinkedIn",
        ])
        ws.append([
            "+971501234567",
            "Omar Khan",
            "",
            "Khan Legal Advisory",
            "Senior Counsel",
            "",
            "Referral",
        ])
        # Width tweaks so the headers don't truncate when the admin opens it.
        widths = {"A": 18, "B": 22, "C": 26, "D": 24, "E": 24, "F": 36, "G": 14}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        # Bold the header row so it's obvious which row is required.
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="contacts_template.xlsx"'
        return response


# --- helpers for ContactViewSet.bulk_upload -------------------------------

_HEADER_ALIASES = {
    "phone": "phone",
    "contact": "phone",
    "contact number": "phone",
    "contact: work phone": "phone",
    "mobile": "phone",
    "name": "full_name",
    "full name": "full_name",
    "deal name": "full_name",
    "email": "email",
    "email address": "email",
    "contact: work e-mail": "email",
    "company": "company",
    "company name": "company",
    "designation": "designation",
    "title": "designation",
    "linkedin": "linkedin_url",
    "linkedin url": "linkedin_url",
    "linkedin profile": "linkedin_url",
    "source": "source",
    "lead source": "source",
}


def _resolve_header_columns(header_row) -> dict[str, int]:
    """Map our canonical field name → column index, given the spreadsheet's
    header row. Tolerates trailing whitespace, mixed case, and the
    common aliases listed in `_HEADER_ALIASES` (so the existing
    LexTalk Excel files would also import without reformatting).
    """
    out: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if not key:
            continue
        canonical = _HEADER_ALIASES.get(key)
        if canonical and canonical not in out:
            out[canonical] = idx
    return out


def _build_source_lookup() -> dict[str, str]:
    """Map every accepted spelling of a Source value (case-folded) → enum value.

    Accepts both the enum value (`WEBSITE_FORM`) and the human label
    (`Website Form`) plus a few aliases callers commonly type.
    """
    lookup = {}
    for value, label in Lead.Source.choices:
        lookup[value.lower()] = value
        lookup[label.lower()] = value
    lookup.update({
        "web": Lead.Source.WEBSITE_FORM,
        "form": Lead.Source.WEBSITE_FORM,
        "linked in": Lead.Source.LINKEDIN,
        "li": Lead.Source.LINKEDIN,
        "ref": Lead.Source.REFERRAL,
        "cold": Lead.Source.COLD_CALL,
    })
    return lookup


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


_PHONE_STRIP_RE = re.compile(r"[\s\-()]")
# After normalization, a "real" phone is an optional leading + followed by
# 7-15 digits (E.164 max). Anything else (letters mid-string, too short, etc.)
# is treated as corrupted and skipped.
_PHONE_VALID_RE = re.compile(r"^\+?\d{7,15}$")


def _looks_like_phone(phone: str) -> bool:
    return bool(_PHONE_VALID_RE.match(phone))


def _normalize_phone(value) -> str:
    """Coerce a cell value into a clean phone string.

    - Strips spaces, hyphens, parens.
    - Float (Excel auto-encodes phone-as-number, e.g. 971552169009.0) → digit string.
    - Returns "" for None / "N/A" / blank → caller treats as missing.
    """
    if value is None:
        return ""
    if isinstance(value, float):
        if value != value or value == 0:  # NaN or 0
            return ""
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if not s or s.lower() in {"n/a", "na", "none", "null", "-"}:
        return ""
    return _PHONE_STRIP_RE.sub("", s)


class SubPipelineViewSet(viewsets.ModelViewSet):
    """CRUD on per-event sub-pipelines.

    - GET (list/retrieve) — any authenticated user (LeadForm dropdown + master filters).
    - POST / PATCH / PUT / DELETE — admin only (the /admin/pipelines CMS page).
    - Filterable by `?event=<id>` and `?is_active=true`.
    - Hard delete is blocked when leads or tiers still reference the row;
      we surface a clean DRF 400 with the "deactivate instead" hint.
    """

    serializer_class = SubPipelineSerializer
    pagination_class = None
    filterset_fields = ("event", "is_active")

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsAuthenticated()]
        return [IsAdmin()]

    def get_queryset(self):
        return SubPipeline.objects.select_related("event").all()

    def destroy(self, request, *args, **kwargs):
        # Same shape as PackageTierViewSet.destroy — convert PROTECT errors
        # into a friendly 400 so the admin page can show the hint inline.
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
        except ProtectedError as e:
            count = len(e.protected_objects)
            raise ValidationError(
                {
                    "detail": (
                        f"Can't delete '{instance.name}' — "
                        f"{count} lead{'s' if count != 1 else ''} or tier{'s' if count != 1 else ''} "
                        f"still attached. Deactivate it instead so it disappears from new-lead "
                        f"dropdowns without breaking existing data."
                    )
                }
            )
        return Response(status=status.HTTP_204_NO_CONTENT)


class PackageTierViewSet(viewsets.ModelViewSet):
    """CRUD on the per-sub-pipeline tier list.

    - GET (list/retrieve) — any authenticated user (LeadForm dropdown needs it).
    - POST / PATCH / PUT / DELETE — admin only (the /admin/tiers CMS page).
    - Filterable by ?sub_pipeline=<id> so the frontend can fetch just
      the tiers relevant to the current sub-pipeline.
    - No pagination — the full list is small (~20 rows) and the dropdown
      needs every active tier in one shot.
    """

    serializer_class = PackageTierSerializer
    pagination_class = None
    filterset_fields = ("sub_pipeline", "is_active")

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsAuthenticated()]
        return [IsAdmin()]

    def get_queryset(self):
        return PackageTier.objects.all()

    def destroy(self, request, *args, **kwargs):
        # PackageTier → Lead is on_delete=PROTECT. Without this catch the
        # raw Django ProtectedError bubbles into a 500 with an HTML debug
        # page; the frontend can't show a useful message. Convert it into
        # a clean 400 so callers see the "deactivate instead" hint.
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
        except ProtectedError as e:
            count = len(e.protected_objects)
            raise ValidationError(
                {
                    "detail": (
                        f"Can't delete '{instance.name}' — "
                        f"{count} lead{'s' if count != 1 else ''} still use this tier. "
                        f"Deactivate it instead so it disappears from the dropdown without breaking existing leads."
                    )
                }
            )
        return Response(status=status.HTTP_204_NO_CONTENT)
